import csv
import json
from pathlib import Path

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QFormLayout, QLineEdit,
    QTableWidget, QTableWidgetItem, QPushButton, QLabel,
    QHeaderView, QAbstractItemView, QMessageBox, QDialog,
    QDialogButtonBox, QMenu, QFileDialog, QComboBox, QCheckBox,
)
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import QFont, QAction

from app.db.database import Database
from app.db.models import SavedWord
from app import styles
from app.i18n import tr


# ---------------------------------------------------------------------------
class _AddWordDialog(QDialog):
    """Modal form to add a word/phrase manually."""

    # (code, translation_key)
    LANG_CHOICES = [
        ("AUTO",  "lang_auto_unknown"),
        ("EN",    "lang_EN"),
        ("TR",    "lang_TR"),
        ("DE",    "lang_DE"),
        ("FR",    "lang_FR"),
        ("ES",    "lang_ES"),
        ("IT",    "lang_IT"),
        ("RU",    "lang_RU"),
        ("JA",    "lang_JA"),
        ("AR",    "lang_AR"),
        ("ZH",    "lang_ZH"),
        ("TR-EN", "lang_TR_EN"),
    ]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle(tr("words_add_dlg_title"))
        self.setModal(True)
        self.setMinimumWidth(400)
        self.setStyleSheet(f"background-color: {styles.C['bg_main']}; color: {styles.C['text']};")

        form = QFormLayout(self)
        form.setContentsMargins(20, 20, 20, 20)
        form.setSpacing(12)

        self._original    = QLineEdit()
        self._original.setPlaceholderText(tr("words_add_ph_orig"))
        self._translation = QLineEdit()
        self._translation.setPlaceholderText(tr("words_add_ph_trans"))

        self._lang = QComboBox()
        for code, key in self.LANG_CHOICES:
            self._lang.addItem(tr(key), code)

        form.addRow(self._label(tr("words_add_label_orig")),  self._original)
        form.addRow(self._label(tr("words_add_label_trans")), self._translation)
        form.addRow(self._label(tr("words_add_label_lang")),  self._lang)

        bb = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel,
            parent=self,
        )
        save_btn   = bb.button(QDialogButtonBox.StandardButton.Save)
        cancel_btn = bb.button(QDialogButtonBox.StandardButton.Cancel)
        if save_btn:
            save_btn.setText(tr("btn_save"))
            save_btn.setStyleSheet(styles.BTN_SUCCESS)
        if cancel_btn:
            cancel_btn.setText(tr("btn_cancel"))
            cancel_btn.setStyleSheet(styles.BTN_GHOST)
        bb.accepted.connect(self._on_accept)
        bb.rejected.connect(self.reject)
        form.addRow(bb)

        self._original.setFocus()

    def _label(self, text: str) -> QLabel:
        lbl = QLabel(text)
        lbl.setStyleSheet(f"color: {styles.C['text_dim']}; background: transparent;")
        return lbl

    def _on_accept(self):
        if not self._original.text().strip() or not self._translation.text().strip():
            QMessageBox.warning(self, tr("words_add_missing_t"),
                                tr("words_add_missing_m"))
            return
        self.accept()

    def data(self) -> tuple[str, str, str]:
        return (
            self._original.text().strip(),
            self._translation.text().strip(),
            self._lang.currentData() or "AUTO",
        )


# ---------------------------------------------------------------------------
class WordsTab(QWidget):
    def __init__(self, db: Database, parent=None):
        super().__init__(parent)
        self._db = db
        # IDs the user has ticked. Kept across search/refresh so filtering
        # doesn't silently throw away the selection.
        self._checked_ids: set[int] = set()
        self._build_ui()
        self._connect_signals()
        self.refresh()

    # ------------------------------------------------------------------
    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 20, 24, 20)
        root.setSpacing(14)

        # Header row
        hdr_row = QHBoxLayout()
        hdr = QLabel(tr("words_header"))
        hdr.setFont(QFont("Segoe UI", 16, QFont.Weight.Bold))
        hdr.setStyleSheet(f"color: {styles.C['accent']}; background: transparent;")
        hdr_row.addWidget(hdr)
        hdr_row.addStretch(1)

        self._count_label = QLabel(tr("words_count", n=0))
        self._count_label.setStyleSheet(
            f"color: {styles.C['text_dim']}; font-size: 12px; background: transparent;"
        )
        hdr_row.addWidget(self._count_label)
        hdr_row.addSpacing(12)

        self._add_btn = QPushButton(tr("words_btn_add"))
        self._add_btn.setStyleSheet(styles.BTN_GHOST)
        self._add_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        hdr_row.addWidget(self._add_btn)

        self._export_btn = QPushButton(tr("words_btn_export"))
        self._export_btn.setStyleSheet(styles.BTN_GHOST)
        self._export_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        hdr_row.addWidget(self._export_btn)
        # Build export menu
        export_menu = QMenu(self._export_btn)
        export_menu.addAction("JSON (.json)").triggered.connect(lambda: self._export("json"))
        export_menu.addAction("CSV (.csv)").triggered.connect(lambda: self._export("csv"))
        export_menu.addAction("Excel (.xlsx)").triggered.connect(lambda: self._export("xlsx"))
        self._export_btn.setMenu(export_menu)

        root.addLayout(hdr_row)

        # Search
        self._search = QLineEdit()
        self._search.setPlaceholderText(tr("words_search_ph"))
        self._search.setMaximumHeight(40)
        root.addWidget(self._search)

        # Bulk-action row
        bulk_row = QHBoxLayout()
        bulk_row.setSpacing(10)

        self._select_all_cb = QCheckBox(tr("words_select_all"))
        self._select_all_cb.setTristate(True)
        self._select_all_cb.setCursor(Qt.CursorShape.PointingHandCursor)
        self._select_all_cb.setStyleSheet(
            f"QCheckBox {{ color: {styles.C['text']}; spacing: 6px; }}"
        )
        bulk_row.addWidget(self._select_all_cb)

        self._bulk_del_btn = QPushButton(tr("words_bulk_del"))
        self._bulk_del_btn.setStyleSheet(styles.BTN_DANGER)
        self._bulk_del_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        self._bulk_del_btn.setEnabled(False)
        bulk_row.addWidget(self._bulk_del_btn)

        bulk_row.addStretch(1)

        self._selection_label = QLabel(tr("words_n_selected", n=0))
        self._selection_label.setStyleSheet(
            f"color: {styles.C['text_dim']}; font-size: 12px; background: transparent;"
        )
        bulk_row.addWidget(self._selection_label)
        root.addLayout(bulk_row)

        # Table  (cols: 0=check, 1=orig, 2=trans, 3=lang, 4=date, 5=delete-btn)
        self._table = QTableWidget(0, 6)
        self._table.setHorizontalHeaderLabels([
            "",
            tr("words_col_original"),
            tr("words_col_translation"),
            tr("words_col_language"),
            tr("words_col_date"),
            "",
        ])
        hh = self._table.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(5, QHeaderView.ResizeMode.Fixed)
        self._table.setColumnWidth(0, 36)
        self._table.setColumnWidth(5, 90)
        self._table.verticalHeader().setVisible(False)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._table.setAlternatingRowColors(False)
        self._table.setShowGrid(False)
        self._table.setStyleSheet(
            f"QTableWidget::item:alternate {{ background-color: {styles.C['bg_panel']}; }}"
        )
        root.addWidget(self._table, 1)

        # Empty state label (shown when no rows)
        self._empty_label = QLabel(tr("words_empty_state"))
        self._empty_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._empty_label.setStyleSheet(
            f"color: {styles.C['text_dim']}; font-size: 15px; background: transparent;"
        )
        self._empty_label.hide()
        root.addWidget(self._empty_label)

    def _connect_signals(self):
        self._search.textChanged.connect(self._on_search)
        self._add_btn.clicked.connect(self._on_add_clicked)
        self._table.itemChanged.connect(self._on_item_changed)
        self._select_all_cb.clicked.connect(self._on_select_all_clicked)
        self._bulk_del_btn.clicked.connect(self._on_bulk_delete)

    # ------------------------------------------------------------------
    def refresh(self):
        words = self._db.get_all()
        self._populate(words)

    def _on_search(self, query: str):
        if query.strip():
            words = self._db.search(query.strip())
        else:
            words = self._db.get_all()
        self._populate(words)

    def _populate(self, words: list[SavedWord]):
        # Block itemChanged while we tear down + rebuild rows so the
        # programmatic CheckState changes don't fire user-style events.
        self._table.blockSignals(True)
        self._table.setRowCount(0)
        if not words:
            self._table.blockSignals(False)
            self._table.hide()
            self._empty_label.show()
            self._count_label.setText(tr("words_count", n=0))
            # Drop selections referring to no-longer-visible rows; the
            # underlying set still persists so search filtering preserves them.
            self._sync_selection_label()
            self._sync_select_all_state()
            return

        self._empty_label.hide()
        self._table.show()
        self._count_label.setText(tr("words_count", n=len(words)))

        for word in words:
            row = self._table.rowCount()
            self._table.insertRow(row)

            # Column 0 — checkbox (stores the word id in UserRole)
            chk = QTableWidgetItem()
            chk.setFlags(Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
            chk.setData(Qt.ItemDataRole.UserRole, word.id)
            chk.setCheckState(
                Qt.CheckState.Checked if word.id in self._checked_ids
                else Qt.CheckState.Unchecked
            )
            self._table.setItem(row, 0, chk)

            self._table.setItem(row, 1, self._cell(word.original_text))
            self._table.setItem(row, 2, self._cell(word.translation))
            self._table.setItem(row, 3, self._cell(word.source_lang))
            self._table.setItem(row, 4, self._cell(word.created_at.strftime("%d.%m.%Y %H:%M")))

            del_btn = QPushButton(tr("btn_delete"))
            del_btn.setStyleSheet(styles.BTN_DANGER)
            del_btn.setFixedHeight(30)
            del_btn.clicked.connect(lambda checked, wid=word.id: self._delete(wid))
            self._table.setCellWidget(row, 5, del_btn)
            self._table.setRowHeight(row, 46)

        self._table.blockSignals(False)
        self._sync_selection_label()
        self._sync_select_all_state()

    def _cell(self, text: str) -> QTableWidgetItem:
        item = QTableWidgetItem(text)
        item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
        return item

    # ------------------------------------------------------------------
    # Selection helpers
    # ------------------------------------------------------------------
    def _on_item_changed(self, item: QTableWidgetItem):
        if item.column() != 0:
            return
        wid = item.data(Qt.ItemDataRole.UserRole)
        if wid is None:
            return
        if item.checkState() == Qt.CheckState.Checked:
            self._checked_ids.add(wid)
        else:
            self._checked_ids.discard(wid)
        self._sync_selection_label()
        self._sync_select_all_state()

    def _visible_row_ids(self) -> list[int]:
        ids = []
        for row in range(self._table.rowCount()):
            it = self._table.item(row, 0)
            if it is not None:
                wid = it.data(Qt.ItemDataRole.UserRole)
                if wid is not None:
                    ids.append(wid)
        return ids

    def _set_all_visible(self, checked: bool):
        self._table.blockSignals(True)
        state = Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked
        for row in range(self._table.rowCount()):
            it = self._table.item(row, 0)
            if it is None:
                continue
            it.setCheckState(state)
            wid = it.data(Qt.ItemDataRole.UserRole)
            if wid is None:
                continue
            if checked:
                self._checked_ids.add(wid)
            else:
                self._checked_ids.discard(wid)
        self._table.blockSignals(False)
        self._sync_selection_label()
        self._sync_select_all_state()

    def _on_select_all_clicked(self):
        # Clicked state toggles via tri-state cycle; we want a clear two-way
        # toggle: any visible checked → uncheck all visible; otherwise check all.
        visible = self._visible_row_ids()
        any_checked = any(wid in self._checked_ids for wid in visible)
        self._set_all_visible(not any_checked)

    def _sync_select_all_state(self):
        visible = self._visible_row_ids()
        if not visible:
            self._select_all_cb.blockSignals(True)
            self._select_all_cb.setCheckState(Qt.CheckState.Unchecked)
            self._select_all_cb.blockSignals(False)
            return
        checked_count = sum(1 for wid in visible if wid in self._checked_ids)
        if checked_count == 0:
            state = Qt.CheckState.Unchecked
        elif checked_count == len(visible):
            state = Qt.CheckState.Checked
        else:
            state = Qt.CheckState.PartiallyChecked
        self._select_all_cb.blockSignals(True)
        self._select_all_cb.setCheckState(state)
        self._select_all_cb.blockSignals(False)

    def _sync_selection_label(self):
        n = len(self._checked_ids)
        self._selection_label.setText(tr("words_n_selected", n=n))
        self._bulk_del_btn.setEnabled(n > 0)
        if n > 0:
            self._bulk_del_btn.setText(tr("words_bulk_del_n", n=n))
            self._export_btn.setText(tr("words_btn_export_sel", n=n))
        else:
            self._bulk_del_btn.setText(tr("words_bulk_del"))
            self._export_btn.setText(tr("words_btn_export"))

    def _delete(self, word_id: int | None):
        if word_id is None:
            return
        reply = QMessageBox.question(
            self, tr("words_confirm_del_t"),
            tr("words_confirm_del_m"),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply == QMessageBox.StandardButton.Yes:
            self._db.delete(word_id)
            self._checked_ids.discard(word_id)
            self._on_search(self._search.text())

    def _on_bulk_delete(self):
        ids = list(self._checked_ids)
        if not ids:
            return
        reply = QMessageBox.question(
            self, tr("words_confirm_bulk_t"),
            tr("words_confirm_bulk_m", n=len(ids)),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        for wid in ids:
            self._db.delete(wid)
        self._checked_ids.clear()
        self._on_search(self._search.text())

    # ------------------------------------------------------------------
    def add_word(self, original: str, translation: str, source_lang: str):
        word = SavedWord(original_text=original, translation=translation, source_lang=source_lang)
        self._db.save_word(word)
        self._on_search(self._search.text())

    # ------------------------------------------------------------------
    def _on_add_clicked(self):
        dlg = _AddWordDialog(self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        original, translation, lang = dlg.data()
        self.add_word(original, translation, lang)

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------
    def _export(self, fmt: str):
        # Selected only if the user has ticked rows; otherwise everything.
        all_words = self._db.get_all()
        if self._checked_ids:
            words = [w for w in all_words if w.id in self._checked_ids]
            scope = tr("words_export_scope_sel")
        else:
            words = all_words
            scope = tr("words_export_scope_all")
        if not words:
            QMessageBox.information(self, tr("words_export_empty_t"),
                                    tr("words_export_empty_m"))
            return

        ext_filter = {
            "json": ("kelimelerim.json", tr("words_export_filter_json")),
            "csv":  ("kelimelerim.csv",  tr("words_export_filter_csv")),
            "xlsx": ("kelimelerim.xlsx", tr("words_export_filter_xlsx")),
        }[fmt]
        default_name, file_filter = ext_filter

        path, _ = QFileDialog.getSaveFileName(
            self, tr("words_export_dlg_title"), default_name, file_filter,
        )
        if not path:
            return

        try:
            if fmt == "json":
                self._write_json(path, words)
            elif fmt == "csv":
                self._write_csv(path, words)
            elif fmt == "xlsx":
                self._write_xlsx(path, words)
        except ImportError as e:
            QMessageBox.warning(
                self, tr("words_export_missing_t"),
                tr("words_export_missing_m", err=str(e)),
            )
            return
        except OSError as e:
            QMessageBox.critical(self, tr("words_export_file_err_t"),
                                 tr("words_export_file_err_m", err=str(e)))
            return

        QMessageBox.information(
            self, tr("words_export_done_t"),
            tr("words_export_done_m", n=len(words), scope=scope, path=path),
        )

    @staticmethod
    def _rows(words: list[SavedWord]) -> list[dict]:
        return [
            {
                "original":    w.original_text,
                "translation": w.translation,
                "language":    w.source_lang,
                "created_at":  w.created_at.isoformat(),
            }
            for w in words
        ]

    def _write_json(self, path: str, words: list[SavedWord]):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(self._rows(words), f, ensure_ascii=False, indent=2)

    def _write_csv(self, path: str, words: list[SavedWord]):
        with open(path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Original", "Translation", "Language", "Created At"])
            for w in words:
                writer.writerow([
                    w.original_text, w.translation, w.source_lang,
                    w.created_at.isoformat(),
                ])

    def _write_xlsx(self, path: str, words: list[SavedWord]):
        from openpyxl import Workbook          # raises ImportError if missing
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Kelimeler"
        ws.append(["Original", "Translation", "Language", "Created At"])

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="5B6AF0")
        for col in ws[1]:
            col.font = header_font
            col.fill = header_fill

        for w in words:
            ws.append([
                w.original_text, w.translation, w.source_lang,
                w.created_at.isoformat(),
            ])

        for column_cells in ws.columns:
            length = max(len(str(c.value or "")) for c in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(60, length + 2)

        wb.save(path)
